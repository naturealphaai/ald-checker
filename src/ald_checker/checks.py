"""All check functions for ALD output validation."""
from __future__ import annotations

import csv
import json
import math
import os
import re
import uuid
from datetime import date
from pathlib import Path

from ald_checker.reference import (
    AREA_UNIT_CONVERSIONS,
    CAPACITY_PLAUSIBILITY,
    CONTINENT_BBOX,
    COUNTRY_ALIASES,
    COUNTRY_BBOX,
    EXTRA_COLUMNS,
    STATUS_ALIASES,
    ALD_COLUMNS,
    VALID_GICS,
    VALID_NATURESENSE,
    VALID_STATUSES,
)


# ── Config loading ────────────────────────────────────────────────────────────

def _load_config() -> dict:
    """Load config from config.toml (package dir, cwd, or ~/.config/ald-checker/)."""
    try:
        import tomllib
    except ImportError:
        try:
            import tomli as tomllib
        except ImportError:
            return {}

    search_paths = [
        Path(__file__).resolve().parent.parent.parent / "config.toml",  # ald-checker repo root (priority)
        Path.home() / ".config" / "ald-checker" / "config.toml",
        Path.cwd() / "ald-checker.toml",  # only if explicitly named
    ]
    for p in search_paths:
        if p.exists():
            with p.open("rb") as f:
                return tomllib.load(f)
    return {}


CONFIG = _load_config()


# ── Helpers ──────────────────────────────────────────────────────────────────

class CheckResult:
    def __init__(self, name: str):
        self.name = name
        self.passed = True
        self.issues: list[str] = []
        self.fixed: list[str] = []
        self.warnings: list[str] = []

    def fail(self, msg: str):
        self.passed = False
        self.issues.append(msg)

    def warn(self, msg: str):
        """Informational warning — doesn't cause failure."""
        self.warnings.append(msg)

    def fix(self, msg: str):
        self.fixed.append(msg)


def _majority_vote(mapping: dict[str, list[int]]) -> tuple[str, float]:
    """Return (winner, confidence) where confidence = winner_count / total."""
    total = sum(len(v) for v in mapping.values())
    winner = max(mapping, key=lambda k: len(mapping[k]))
    return winner, len(mapping[winner]) / total


def _smart_title_case(s: str) -> str:
    """Title-case preserving acronyms and special patterns.

    Only converts words that are ALL-CAPS (>3 chars) or all-lowercase.
    Preserves: "A/S", "LLC", "HQ", "II", "#102", "McDonald's".
    """
    words = s.split()
    result = []
    for w in words:
        if len(w) <= 3 and w == w.upper():
            result.append(w)
        elif w != w.upper() and w != w.lower():
            result.append(w)
        elif w[0] in "#0123456789":
            result.append(w)
        elif w == w.upper() and len(w) > 3:
            result.append(w.title())
        elif w == w.lower() and len(w) > 1:
            result.append(w.title())
        else:
            result.append(w)
    return " ".join(result)


def _try_llm_import():
    """Import LLM module, returning None if litellm not installed."""
    try:
        from ald_checker import llm
        return llm
    except ImportError:
        return None


# ── Check functions ──────────────────────────────────────────────────────────

COLUMN_ALIASES = {
    "location": "address",
    "addr": "address",
    "full_address": "address",
    "lat": "latitude",
    "lng": "longitude",
    "lon": "longitude",
    "long": "longitude",
    "type": "asset_type_raw",
    "asset_type": "asset_type_raw",
    "raw_type": "asset_type_raw",
    "naturesense": "naturesense_asset_type",
    "ns_type": "naturesense_asset_type",
    "gics": "industry_code",
    "gics_code": "industry_code",
    "isin": "entity_isin",
    "parent": "parent_name",
    "stake": "entity_stake_pct",
    "stake_pct": "entity_stake_pct",
    "source": "attribution_source",
    "date": "date_researched",
    "details": "supplementary_details",
    "url": "source_url",
}


def check_columns(rows: list[dict], headers: list[str], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """All expected ALD columns exist. Fix renames aliases, drops blanks, adds missing."""
    result = CheckResult("all_columns_exist")

    if fix:
        # Remove blank column names
        blank_cols = [h for h in headers if not h or not h.strip()]
        if blank_cols:
            for row in rows:
                for blank in blank_cols:
                    row.pop(blank, None)
            headers[:] = [h for h in headers if h and h.strip()]
            result.fix(f"Removed {len(blank_cols)} blank column(s)")

        # Rename aliases
        for old_name, new_name in COLUMN_ALIASES.items():
            if old_name in headers and new_name not in headers:
                idx = headers.index(old_name)
                headers[idx] = new_name
                for row in rows:
                    if old_name in row:
                        row[new_name] = row.pop(old_name)
                result.fix(f"Renamed column '{old_name}' → '{new_name}'")

        # Add missing columns with empty values
        missing = [c for c in ALD_COLUMNS if c not in headers]
        for col in missing:
            headers.append(col)
            for row in rows:
                row[col] = ""
            result.fix(f"Added missing column '{col}'")

    missing = [c for c in ALD_COLUMNS if c not in headers]
    extra = [c for c in headers if c not in ALD_COLUMNS and c not in EXTRA_COLUMNS]
    if missing:
        result.fail(f"Missing columns: {missing}")
    if extra:
        result.fail(f"Unexpected columns: {extra}")
    return result


def _load_type_mappings() -> dict[str, dict]:
    """Load asset_type_raw → (NS, GICS) mappings from corp-graph."""
    try:
        import psycopg
        from psycopg.rows import dict_row
        import os
        db_url = os.environ.get("CORPGRAPH_DB_URL", "postgresql://corpgraph:corpgraph@localhost:5432/corpgraph")
        conn = psycopg.connect(db_url, row_factory=dict_row)
        cur = conn.execute("SELECT raw_type, naturesense_asset_type, industry_code FROM asset_type_mappings")
        mappings = {row["raw_type"]: {"ns": row["naturesense_asset_type"], "gics": row["industry_code"]} for row in cur.fetchall()}
        conn.close()
        return mappings
    except Exception:
        return {}


def _save_type_mapping(raw_type: str, ns: str, gics: str, source: str = "llm"):
    """Save a new mapping back to corp-graph."""
    try:
        import psycopg
        import os
        db_url = os.environ.get("CORPGRAPH_DB_URL", "postgresql://corpgraph:corpgraph@localhost:5432/corpgraph")
        conn = psycopg.connect(db_url, autocommit=True)
        conn.execute(
            "INSERT INTO asset_type_mappings (raw_type, naturesense_asset_type, industry_code, source) "
            "VALUES (%s, %s, %s, %s) ON CONFLICT (raw_type) DO UPDATE SET "
            "naturesense_asset_type = EXCLUDED.naturesense_asset_type, "
            "industry_code = EXCLUDED.industry_code, updated_at = NOW()",
            (raw_type, ns, gics, source),
        )
        conn.close()
    except Exception:
        pass


def check_none_strings(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """Convert literal 'None', 'null', 'N/A', 'nan' strings to empty across all columns."""
    result = CheckResult("none_strings")
    NONE_VALUES = {"None", "null", "N/A", "nan", "NaN", "none", "NULL", "n/a", "NA"}
    count = 0
    if fix:
        for row in rows:
            for key in row:
                if row[key] in NONE_VALUES:
                    row[key] = ""
                    count += 1
        if count:
            result.fix(f"Cleared {count} None/null/N-A strings across all columns")
    else:
        for row in rows:
            for key in row:
                if row[key] in NONE_VALUES:
                    count += 1
        if count:
            result.fail(f"{count} None/null/N-A strings found across all columns")
    return result


def check_numeric_cleanup(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """Strip '.0' suffix from numeric string columns (industry_code, entity_stake_pct)."""
    result = CheckResult("numeric_cleanup")
    NUMERIC_COLS = ["industry_code", "entity_stake_pct"]
    count = 0
    for row in rows:
        for col in NUMERIC_COLS:
            val = row.get(col, "")
            if isinstance(val, str) and val.endswith(".0"):
                if fix:
                    row[col] = val[:-2]
                count += 1
            elif isinstance(val, float):
                if fix:
                    row[col] = str(int(val)) if val == int(val) else str(val)
                count += 1
    if fix and count:
        result.fix(f"Cleaned {count} numeric values (stripped .0 suffix)")
    elif count:
        result.fail(f"{count} numeric values have .0 suffix")
    return result


def check_asset_type_raw_standardize(rows: list[dict], fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """Standardize asset_type_raw values — fix typos, normalize format."""
    result = CheckResult("asset_type_raw_standardize")

    # Get unique raw types
    raw_types: dict[str, list[int]] = {}
    for i, row in enumerate(rows):
        raw = row.get("asset_type_raw", "").strip()
        if raw:
            raw_types.setdefault(raw, []).append(i)

    if not fix_llm:
        return result  # Nothing to check in non-fix mode

    # Load known mappings — if a raw type (lowered) is already known, standardize to that form
    mappings = _load_type_mappings()
    known_lower = {k.lower(): k for k in mappings}

    still_unknown = {}
    for raw, idxs in raw_types.items():
        if raw.lower() in known_lower:
            canonical = known_lower[raw.lower()]
            if raw != canonical:
                for idx in idxs:
                    rows[idx]["asset_type_raw"] = canonical
                result.fix(f"'{raw}' → '{canonical}' at {len(idxs)} rows")
        else:
            still_unknown[raw] = idxs

    # LLM standardize unknown types
    if still_unknown:
        llm = _try_llm_import()
        if llm:
            unique_types = list(still_unknown.keys())
            try:
                standardized = llm.standardize_raw_types(unique_types, model=model)
                for original, cleaned in standardized.items():
                    if original in still_unknown and cleaned != original:
                        for idx in still_unknown[original]:
                            rows[idx]["asset_type_raw"] = cleaned
                        result.fix(f"LLM: '{original}' → '{cleaned}' at {len(still_unknown[original])} rows")
            except Exception as e:
                result.fail(f"LLM standardization failed: {e}")

    return result


def check_naturesense_correct(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """NatureSense is correct for the asset_type_raw (not just valid)."""
    result = CheckResult("naturesense_correct")
    mappings = _load_type_mappings()

    wrong: dict[str, list[int]] = {}  # (raw, current_ns, expected_ns) -> idxs
    unknown_raw: dict[str, list[int]] = {}

    missing_ns: dict[str, list[int]] = {}

    for i, row in enumerate(rows):
        raw = row.get("asset_type_raw", "").strip().lower()
        if not raw:
            continue
        ns = row.get("naturesense_asset_type", "").strip()
        is_empty = not ns or ns == "None"

        if raw in mappings:
            expected_ns = mappings[raw]["ns"]
            if is_empty:
                key = (raw, "", expected_ns)
                wrong.setdefault(key, []).append(i)
            elif ns != expected_ns:
                key = (raw, ns, expected_ns)
                wrong.setdefault(key, []).append(i)
        else:
            if is_empty:
                missing_ns.setdefault(raw, []).append(i)
            else:
                unknown_raw.setdefault(raw, []).append(i)

    if not fix and not fix_llm:
        for (raw, current, expected), idxs in wrong.items():
            result.fail(f"'{raw}': NS is '{current}' but should be '{expected}' at {len(idxs)} rows")
        return result

    # Fix known wrong mappings
    for (raw, current, expected), idxs in wrong.items():
        for idx in idxs:
            rows[idx]["naturesense_asset_type"] = expected
        result.fix(f"'{raw}': '{current}' → '{expected}' at {len(idxs)} rows")

    # Merge missing_ns into unknown_raw for LLM
    for raw, idxs in missing_ns.items():
        unknown_raw.setdefault(raw, []).extend(idxs)

    # For unknown raw types, use LLM to determine correct NS and save to corp-graph
    if unknown_raw and fix_llm:
        llm = _try_llm_import()
        if llm:
            try:
                raw_list = list(unknown_raw.keys())
                ns_results = llm.classify_naturesense(raw_list, model=model)
                for raw, ns_result in ns_results.items():
                    if raw in unknown_raw and ns_result in VALID_NATURESENSE:
                        for idx in unknown_raw[raw]:
                            rows[idx]["naturesense_asset_type"] = ns_result
                        result.fix(f"LLM: '{raw}' → NS='{ns_result}' at {len(unknown_raw[raw])} rows")
                        # Don't save to mapping table here — GICS hasn't been classified yet
                        # check_gics_correct will save the full mapping after GICS is determined
            except Exception:
                pass

    return result


def check_gics_correct(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """GICS is correct for the asset_type_raw (not just valid)."""
    result = CheckResult("gics_correct")
    mappings = _load_type_mappings()

    wrong: dict[str, list[int]] = {}
    unknown_raw: dict[str, list[int]] = {}

    missing_gics: dict[str, list[int]] = {}

    for i, row in enumerate(rows):
        raw = row.get("asset_type_raw", "").strip().lower()
        if not raw:
            continue
        gics = str(row.get("industry_code", "")).strip()
        is_empty = not gics or gics == "None" or gics == ""

        if raw in mappings:
            expected_gics = mappings[raw]["gics"]
            if is_empty:
                # Empty GICS but we know what it should be
                key = (raw, "", expected_gics)
                wrong.setdefault(key, []).append(i)
            elif gics != expected_gics:
                key = (raw, gics, expected_gics)
                wrong.setdefault(key, []).append(i)
        else:
            if is_empty:
                missing_gics.setdefault(raw, []).append(i)
            else:
                unknown_raw.setdefault(raw, []).append(i)

    if not fix and not fix_llm:
        for (raw, current, expected), idxs in wrong.items():
            result.fail(f"'{raw}': GICS is '{current}' but should be '{expected}' at {len(idxs)} rows")
        return result

    # Fix known wrong
    for (raw, current, expected), idxs in wrong.items():
        for idx in idxs:
            rows[idx]["industry_code"] = expected
        result.fix(f"'{raw}': GICS '{current}' → '{expected}' at {len(idxs)} rows")

    # Merge missing_gics into unknown_raw for LLM classification
    for raw, idxs in missing_gics.items():
        unknown_raw.setdefault(raw, []).extend(idxs)

    # LLM for unknown
    if unknown_raw and fix_llm:
        llm = _try_llm_import()
        if llm:
            try:
                raw_list = list(unknown_raw.keys())
                gics_results = llm.classify_gics(raw_list, model=model)
                for raw, gics_result in gics_results.items():
                    if raw in unknown_raw and gics_result in VALID_GICS:
                        for idx in unknown_raw[raw]:
                            rows[idx]["industry_code"] = gics_result
                        result.fix(f"LLM: '{raw}' → GICS='{gics_result}' at {len(unknown_raw[raw])} rows")
                        ns = rows[unknown_raw[raw][0]].get("naturesense_asset_type", "")
                        if ns:
                            _save_type_mapping(raw, ns, gics_result, "llm")
            except Exception:
                pass

    return result


def _load_asset_id_registry() -> dict[str, dict]:
    """Load asset_id registry from corp-graph. Returns {asset_id: {source, entity_isin}}."""
    try:
        import psycopg
        from psycopg.rows import dict_row
        import os
        db_url = os.environ.get("CORPGRAPH_DB_URL", "postgresql://corpgraph:corpgraph@localhost:5432/corpgraph")
        conn = psycopg.connect(db_url, row_factory=dict_row)
        cur = conn.execute("SELECT asset_id, source, entity_isin FROM asset_id_registry")
        registry = {row["asset_id"]: {"source": row["source"], "isin": row.get("entity_isin", "")} for row in cur.fetchall()}
        conn.close()
        return registry
    except Exception:
        return {}


def _generate_unique_id(registry: dict) -> str:
    """Generate a UUID that doesn't collide with the registry."""
    for _ in range(10):
        new_id = str(uuid.uuid4())
        if new_id not in registry:
            return new_id
    return str(uuid.uuid4())  # Astronomically unlikely to still collide


def _coords_similar(row_a: dict, isin_b: str, registry_entry: dict) -> bool:
    """Check if a row likely matches a registry entry (same company)."""
    row_isin = row_a.get("entity_isin", "")
    reg_isin = registry_entry.get("isin", "")
    if row_isin and reg_isin and row_isin == reg_isin:
        return True
    return False


def check_asset_id_unique(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """All asset_ids are unique, non-empty UUIDs. Cross-checks against corp-graph registry."""
    result = CheckResult("asset_id_unique")
    registry = _load_asset_id_registry()
    seen: dict[str, int] = {}
    empty_rows = []

    for i, row in enumerate(rows):
        aid = row.get("asset_id", "").strip()
        if not aid:
            empty_rows.append(i)
            if fix:
                row["asset_id"] = _generate_unique_id(registry)
                result.fix(f"Row {i}: generated asset_id {row['asset_id']}")
            continue

        # Check for in-file duplicates
        if aid in seen:
            if fix:
                row["asset_id"] = _generate_unique_id(registry)
                result.fix(f"Row {i}: regenerated duplicate asset_id (was '{aid}')")
            else:
                result.fail(f"Duplicate asset_id '{aid}' at rows {seen[aid]} and {i}")
            continue

        seen[aid] = i

        # Cross-check against registry
        if aid in registry:
            reg = registry[aid]
            row_isin = row.get("entity_isin", "")
            reg_isin = reg.get("isin", "")

            if row_isin and reg_isin and row_isin != reg_isin:
                # ID belongs to a DIFFERENT company
                if fix:
                    row["asset_id"] = _generate_unique_id(registry)
                    result.fix(f"Row {i}: asset_id '{aid}' belongs to {reg_isin}, regenerated for {row_isin}")
                else:
                    result.fail(f"Row {i}: asset_id '{aid}' exists in registry for {reg_isin}, but this row is {row_isin}")
            # Same company or can't tell → allow (carried-over ALD ID)

    if empty_rows and not fix:
        result.fail(f"Empty asset_id at {len(empty_rows)} rows: {empty_rows[:10]}{'...' if len(empty_rows) > 10 else ''}")

    for i, row in enumerate(rows):
        aid = row.get("asset_id", "").strip()
        if aid:
            try:
                uuid.UUID(aid)
            except ValueError:
                result.fail(f"Row {i}: asset_id '{aid}' is not a valid UUID")

    return result


def check_naturesense_valid(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """All naturesense_asset_type values are in the official list."""
    result = CheckResult("naturesense_valid")
    invalid: dict[str, list[int]] = {}
    for i, row in enumerate(rows):
        ns = row.get("naturesense_asset_type", "").strip()
        if ns and ns not in VALID_NATURESENSE:
            invalid.setdefault(ns, []).append(i)

    if not fix and not fix_llm:
        for ns, idxs in invalid.items():
            result.fail(f"Invalid naturesense_asset_type '{ns}' at {len(idxs)} rows (first: {idxs[:5]})")
        return result

    # Deterministic fixes first
    still_invalid: dict[str, list[int]] = {}
    for ns, idxs in invalid.items():
        if ns.lower().startswith("other"):
            canonical = "Other (5km buffer area of influence)"
            for idx in idxs:
                rows[idx]["naturesense_asset_type"] = canonical
            result.fix(f"'{ns}' → '{canonical}' at {len(idxs)} rows")
        else:
            ns_lower = ns.lower()
            match = next((v for v in VALID_NATURESENSE if v.lower() == ns_lower), None)
            if match:
                for idx in idxs:
                    rows[idx]["naturesense_asset_type"] = match
                result.fix(f"Case fix '{ns}' → '{match}' at {len(idxs)} rows")
            else:
                still_invalid[ns] = idxs

    # LLM fix for remaining
    if still_invalid and fix_llm:
        llm = _try_llm_import()
        if llm:
            raw_for_ns: dict[str, str] = {}
            for ns, idxs in still_invalid.items():
                raws: dict[str, int] = {}
                for idx in idxs:
                    r = rows[idx].get("asset_type_raw", "").strip()
                    if r:
                        raws[r] = raws.get(r, 0) + 1
                raw_for_ns[ns] = max(raws, key=raws.get) if raws else ns

            try:
                llm_map = llm.classify_naturesense(list(raw_for_ns.values()), model or llm.DEFAULT_MODEL)
                for ns, idxs in still_invalid.items():
                    classified = llm_map.get(raw_for_ns[ns], "")
                    if classified in VALID_NATURESENSE:
                        for idx in idxs:
                            rows[idx]["naturesense_asset_type"] = classified
                        result.fix(f"LLM: '{ns}' → '{classified}' at {len(idxs)} rows")
                    else:
                        result.fail(f"LLM returned invalid '{classified}' for '{ns}' — skipped")
            except Exception as e:
                result.fail(f"LLM classification failed: {e}")
                for ns, idxs in still_invalid.items():
                    result.fail(f"Invalid naturesense_asset_type '{ns}' at {len(idxs)} rows")
        else:
            result.fail("--fix-llm requires litellm: pip install ald-checker[llm]")
    elif still_invalid:
        for ns, idxs in still_invalid.items():
            result.fail(f"Invalid naturesense_asset_type '{ns}' at {len(idxs)} rows (first: {idxs[:5]})")

    return result


def check_naturesense_consistency(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """Same asset_type_raw always maps to the same naturesense_asset_type."""
    result = CheckResult("naturesense_consistency")
    mapping: dict[str, dict[str, list[int]]] = {}
    for i, row in enumerate(rows):
        raw = row.get("asset_type_raw", "").strip().lower()
        ns = row.get("naturesense_asset_type", "").strip()
        if raw and ns:
            mapping.setdefault(raw, {}).setdefault(ns, []).append(i)

    inconsistent = {raw: ns_map for raw, ns_map in mapping.items() if len(ns_map) > 1}
    if not inconsistent:
        return result

    if not fix and not fix_llm:
        for raw, ns_map in inconsistent.items():
            counts = {ns: len(idxs) for ns, idxs in ns_map.items()}
            result.fail(f"'{raw}' maps to multiple naturesense types: {counts}")
        return result

    needs_llm = []
    for raw, ns_map in inconsistent.items():
        winner, confidence = _majority_vote(ns_map)
        counts = {ns: len(idxs) for ns, idxs in ns_map.items()}

        if confidence >= 0.7:
            fixed_count = 0
            for ns, idxs in ns_map.items():
                if ns != winner:
                    for idx in idxs:
                        rows[idx]["naturesense_asset_type"] = winner
                    fixed_count += len(idxs)
            result.fix(f"'{raw}': majority-voted → '{winner}' (fixed {fixed_count} rows, was {counts})")
        elif fix_llm:
            needs_llm.append(raw)
        else:
            fixed_count = 0
            for ns, idxs in ns_map.items():
                if ns != winner:
                    for idx in idxs:
                        rows[idx]["naturesense_asset_type"] = winner
                    fixed_count += len(idxs)
            result.fix(f"'{raw}': majority-voted → '{winner}' (fixed {fixed_count} rows, was {counts}, low confidence {confidence:.0%})")

    if needs_llm:
        llm = _try_llm_import()
        if llm:
            try:
                llm_map = llm.classify_naturesense(needs_llm, model or llm.DEFAULT_MODEL)
                for raw in needs_llm:
                    ns_map = inconsistent[raw]
                    classified = llm_map.get(raw, "")
                    counts = {ns: len(idxs) for ns, idxs in ns_map.items()}
                    if classified in VALID_NATURESENSE:
                        fixed_count = 0
                        for ns, idxs in ns_map.items():
                            if ns != classified:
                                for idx in idxs:
                                    rows[idx]["naturesense_asset_type"] = classified
                                fixed_count += len(idxs)
                        result.fix(f"'{raw}': LLM classified → '{classified}' (fixed {fixed_count} rows, was {counts})")
                    else:
                        winner, _ = _majority_vote(ns_map)
                        fixed_count = 0
                        for ns, idxs in ns_map.items():
                            if ns != winner:
                                for idx in idxs:
                                    rows[idx]["naturesense_asset_type"] = winner
                                fixed_count += len(idxs)
                        result.fix(f"'{raw}': LLM returned invalid '{classified}', fell back to majority → '{winner}' (fixed {fixed_count} rows)")
            except Exception as e:
                result.fail(f"LLM classification failed: {e}")
                for raw in needs_llm:
                    ns_map = inconsistent[raw]
                    winner, _ = _majority_vote(ns_map)
                    counts = {ns: len(idxs) for ns, idxs in ns_map.items()}
                    fixed_count = 0
                    for ns, idxs in ns_map.items():
                        if ns != winner:
                            for idx in idxs:
                                rows[idx]["naturesense_asset_type"] = winner
                            fixed_count += len(idxs)
                    result.fix(f"'{raw}': LLM failed, majority-voted → '{winner}' (fixed {fixed_count} rows, was {counts})")
        else:
            result.fail("--fix-llm requires litellm: pip install ald-checker[llm]")

    return result


def check_gics_valid(rows: list[dict], fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """All industry_code values are valid 6-digit GICS codes."""
    result = CheckResult("gics_valid")
    invalid: dict[str, list[int]] = {}
    for i, row in enumerate(rows):
        code = row.get("industry_code", "").strip()
        if code and code not in VALID_GICS:
            invalid.setdefault(code, []).append(i)

    if not invalid:
        return result

    if fix_llm:
        llm = _try_llm_import()
        if llm:
            raw_for_code: dict[str, set[str]] = {}
            for code, idxs in invalid.items():
                raws = set()
                for idx in idxs:
                    r = rows[idx].get("asset_type_raw", "").strip()
                    if r:
                        raws.add(r)
                raw_for_code[code] = raws

            all_raws = list({r for raws in raw_for_code.values() for r in raws})
            if all_raws:
                try:
                    llm_map = llm.classify_gics(all_raws, model or llm.DEFAULT_MODEL)
                    for code, idxs in invalid.items():
                        for idx in idxs:
                            raw = rows[idx].get("asset_type_raw", "").strip()
                            new_code = llm_map.get(raw, "")
                            if new_code in VALID_GICS:
                                rows[idx]["industry_code"] = new_code
                            else:
                                result.fail(f"Row {idx}: LLM returned invalid GICS '{new_code}' for '{raw}'")
                        raws_str = ", ".join(raw_for_code[code])
                        result.fix(f"LLM: invalid code '{code}' reclassified for '{raws_str}' at {len(idxs)} rows")
                except Exception as e:
                    result.fail(f"LLM GICS classification failed: {e}")
                    for code, idxs in invalid.items():
                        result.fail(f"Invalid GICS code '{code}' at {len(idxs)} rows (first: {idxs[:5]})")
            else:
                for code, idxs in invalid.items():
                    result.fail(f"Invalid GICS code '{code}' at {len(idxs)} rows (first: {idxs[:5]})")
        else:
            result.fail("--fix-llm requires litellm: pip install ald-checker[llm]")
    else:
        for code, idxs in invalid.items():
            result.fail(f"Invalid GICS code '{code}' at {len(idxs)} rows (first: {idxs[:5]})")

    return result


def check_gics_consistency(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """Same asset_type_raw always maps to the same industry_code."""
    result = CheckResult("gics_consistency")
    mapping: dict[str, dict[str, list[int]]] = {}
    for i, row in enumerate(rows):
        raw = row.get("asset_type_raw", "").strip().lower()
        code = row.get("industry_code", "").strip()
        if raw and code:
            mapping.setdefault(raw, {}).setdefault(code, []).append(i)

    inconsistent = {raw: code_map for raw, code_map in mapping.items() if len(code_map) > 1}
    if not inconsistent:
        return result

    if not fix and not fix_llm:
        for raw, code_map in inconsistent.items():
            counts = {code: len(idxs) for code, idxs in code_map.items()}
            result.fail(f"'{raw}' maps to multiple GICS codes: {counts}")
        return result

    needs_llm = []
    for raw, code_map in inconsistent.items():
        winner, confidence = _majority_vote(code_map)
        counts = {code: len(idxs) for code, idxs in code_map.items()}

        if confidence >= 0.7:
            fixed_count = 0
            for code, idxs in code_map.items():
                if code != winner:
                    for idx in idxs:
                        rows[idx]["industry_code"] = winner
                    fixed_count += len(idxs)
            result.fix(f"'{raw}': majority-voted → '{winner}' (fixed {fixed_count} rows, was {counts})")
        elif fix_llm:
            needs_llm.append(raw)
        else:
            fixed_count = 0
            for code, idxs in code_map.items():
                if code != winner:
                    for idx in idxs:
                        rows[idx]["industry_code"] = winner
                    fixed_count += len(idxs)
            result.fix(f"'{raw}': majority-voted → '{winner}' (fixed {fixed_count} rows, was {counts}, low confidence {confidence:.0%})")

    if needs_llm:
        llm = _try_llm_import()
        if llm:
            try:
                llm_map = llm.classify_gics(needs_llm, model or llm.DEFAULT_MODEL)
                for raw in needs_llm:
                    code_map = inconsistent[raw]
                    classified = llm_map.get(raw, "")
                    counts = {code: len(idxs) for code, idxs in code_map.items()}
                    if classified in VALID_GICS:
                        fixed_count = 0
                        for code, idxs in code_map.items():
                            if code != classified:
                                for idx in idxs:
                                    rows[idx]["industry_code"] = classified
                                fixed_count += len(idxs)
                        result.fix(f"'{raw}': LLM classified → '{classified}' (fixed {fixed_count} rows, was {counts})")
                    else:
                        winner, _ = _majority_vote(code_map)
                        fixed_count = 0
                        for code, idxs in code_map.items():
                            if code != winner:
                                for idx in idxs:
                                    rows[idx]["industry_code"] = winner
                                fixed_count += len(idxs)
                        result.fix(f"'{raw}': LLM returned invalid '{classified}', fell back to majority → '{winner}' (fixed {fixed_count} rows)")
            except Exception as e:
                result.fail(f"LLM GICS classification failed: {e}")
                for raw in needs_llm:
                    code_map = inconsistent[raw]
                    winner, _ = _majority_vote(code_map)
                    counts = {code: len(idxs) for code, idxs in code_map.items()}
                    fixed_count = 0
                    for code, idxs in code_map.items():
                        if code != winner:
                            for idx in idxs:
                                rows[idx]["industry_code"] = winner
                            fixed_count += len(idxs)
                    result.fix(f"'{raw}': LLM failed, majority-voted → '{winner}' (fixed {fixed_count} rows, was {counts})")
        else:
            result.fail("--fix-llm requires litellm: pip install ald-checker[llm]")

    return result


def check_coordinates(rows: list[dict], **_kw) -> CheckResult:
    """Lat/lon in valid ranges, not at null island, not swapped, not in ocean."""
    result = CheckResult("coordinates")

    # Known ocean regions (no land) — conservative boxes
    OCEAN_BOXES = [
        # Mid-Pacific (no islands)
        (10, 40, -170, -130, "mid-Pacific Ocean"),
        # South Pacific
        (-60, -30, -170, -90, "South Pacific Ocean"),
        # Mid-Atlantic (no islands in this band)
        (10, 35, -60, -20, "mid-Atlantic Ocean"),
        # South Atlantic
        (-55, -20, -40, 0, "South Atlantic Ocean"),
        # Southern Indian Ocean
        (-60, -35, 30, 100, "Southern Indian Ocean"),
        # Arctic Ocean
        (80, 90, -180, 180, "Arctic Ocean"),
        # Antarctic
        (-90, -70, -180, 180, "Antarctica/Southern Ocean"),
    ]

    for i, row in enumerate(rows):
        lat_s = row.get("latitude", "").strip()
        lon_s = row.get("longitude", "").strip()
        name = row.get("name", "").strip()
        if not lat_s or not lon_s:
            continue
        try:
            lat, lon = float(lat_s), float(lon_s)
        except ValueError:
            result.fail(f"Row {i} '{name}': non-numeric coordinates lat='{lat_s}' lon='{lon_s}'")
            continue
        if not (-90 <= lat <= 90):
            result.fail(f"Row {i} '{name}': latitude {lat} out of range [-90, 90]")
        if not (-180 <= lon <= 180):
            result.fail(f"Row {i} '{name}': longitude {lon} out of range [-180, 180]")
        if abs(lat) < 0.01 and abs(lon) < 0.01:
            result.fail(f"Row {i} '{name}': coordinates ({lat}, {lon}) suspiciously near null island")
        if (-90 <= lon <= 90) and not (-90 <= lat <= 90) and (-180 <= lat <= 180):
            result.fail(f"Row {i} '{name}': lat={lat}, lon={lon} — possibly swapped?")

        # Ocean box check (known deep-ocean regions, no land)
        for min_lat, max_lat, min_lon, max_lon, region in OCEAN_BOXES:
            if min_lat <= lat <= max_lat and min_lon <= lon <= max_lon:
                result.fail(f"Row {i} '{name}': ({lat}, {lon}) is in the {region}")
                break

    # Land/water check using global_land_mask (if installed)
    try:
        from global_land_mask import globe
        water_rows = []
        for i, row in enumerate(rows):
            lat_s = row.get("latitude", "").strip()
            lon_s = row.get("longitude", "").strip()
            name = row.get("name", "").strip()
            atype = row.get("asset_type_raw", "").strip().lower()
            if not lat_s or not lon_s:
                continue
            try:
                lat, lon = float(lat_s), float(lon_s)
            except ValueError:
                continue
            if not globe.is_land(lat, lon):
                # Skip known offshore asset types
                offshore_types = {"offshore wind farm", "offshore platform", "oil platform",
                                  "subsea", "fpso", "offshore", "wind farm"}
                if any(ot in atype for ot in offshore_types):
                    continue
                water_rows.append((i, lat, lon, name))
        if water_rows:
            for i, lat, lon, name in water_rows:
                result.warn(
                    f"Row {i} '{name}': ({lat}, {lon}) appears to be in water "
                    f"(may be coastal, island, or offshore — verify)"
                )
    except ImportError:
        pass  # global_land_mask not installed, skip

    return result


def _try_reverse_geocoder():
    """Try to import reverse_geocoder. Returns module or None."""
    try:
        import reverse_geocoder as rg
        return rg
    except ImportError:
        return None


def _parse_country_from_address(address: str) -> str | None:
    """Extract country from the last segment of an address string."""
    if not address:
        return None
    parts = [p.strip() for p in address.split(",")]
    if not parts:
        return None
    candidate = parts[-1].lower().strip()
    candidate = re.sub(r"^\d[\d\s-]*", "", candidate).strip()
    candidate = re.sub(r"\d[\d\s-]*$", "", candidate).strip()
    if not candidate:
        return None
    if candidate in COUNTRY_ALIASES:
        return COUNTRY_ALIASES[candidate]
    if candidate in COUNTRY_BBOX:
        return candidate
    for country in COUNTRY_BBOX:
        if country in candidate:
            return country
    return None


def check_coords_country_match(rows: list[dict], **_kw) -> CheckResult:
    """Coordinates are in the correct country based on the address field.

    Uses reverse_geocoder (offline KD-tree, 100K+ cities) if installed,
    falls back to country bounding boxes. Warn-only.
    """
    result = CheckResult("coords_country_match")

    # Collect rows with both address and coords
    to_check = []
    for i, row in enumerate(rows):
        address = row.get("address", "").strip()
        lat_s = row.get("latitude", "").strip()
        lon_s = row.get("longitude", "").strip()
        name = row.get("name", "").strip()
        if not address or not lat_s or not lon_s:
            continue
        try:
            lat, lon = float(lat_s), float(lon_s)
        except ValueError:
            continue
        addr_country = _parse_country_from_address(address)
        if addr_country:
            to_check.append((i, lat, lon, name, address, addr_country))

    if not to_check:
        return result

    rg = _try_reverse_geocoder()
    if rg:
        # Batch reverse geocode all coords at once (fast)
        coords = [(lat, lon) for _, lat, lon, _, _, _ in to_check]
        geo_results = rg.search(coords)
        for (idx, lat, lon, name, address, addr_country), geo in zip(to_check, geo_results):
            coord_cc = geo["cc"].lower()
            # Map country name to CC for comparison
            addr_cc = _country_to_cc(addr_country)
            if addr_cc and coord_cc != addr_cc:
                result.warn(
                    f"Row {idx} '{name}': coords ({lat}, {lon}) → "
                    f"{geo['name']}, {geo['cc']} but address says {addr_country.title()}"
                )
    else:
        # Fallback: bounding box check
        for idx, lat, lon, name, address, addr_country in to_check:
            if addr_country not in COUNTRY_BBOX:
                continue
            min_lat, max_lat, min_lon, max_lon = COUNTRY_BBOX[addr_country]
            if not (min_lat <= lat <= max_lat and min_lon <= lon <= max_lon):
                result.warn(
                    f"Row {idx} '{name}': coords ({lat}, {lon}) outside "
                    f"{addr_country.title()} (address: '{address[-60:]}')"
                )
    return result


def _country_to_cc(country: str) -> str | None:
    """Map country name to ISO 2-letter code."""
    mapping = {
        "taiwan": "tw", "japan": "jp", "china": "cn", "south korea": "kr",
        "korea": "kr", "india": "in", "singapore": "sg", "vietnam": "vn",
        "indonesia": "id", "thailand": "th", "malaysia": "my", "philippines": "ph",
        "germany": "de", "netherlands": "nl", "france": "fr",
        "united kingdom": "gb", "uk": "gb", "ireland": "ie",
        "italy": "it", "spain": "es", "switzerland": "ch",
        "sweden": "se", "norway": "no", "denmark": "dk", "finland": "fi",
        "belgium": "be", "austria": "at", "poland": "pl",
        "czech republic": "cz", "portugal": "pt",
        "united states": "us", "us": "us", "usa": "us",
        "canada": "ca", "mexico": "mx", "brazil": "br",
        "argentina": "ar", "chile": "cl", "colombia": "co",
        "australia": "au", "new zealand": "nz",
        "saudi arabia": "sa", "uae": "ae", "israel": "il",
        "south africa": "za", "egypt": "eg", "turkey": "tr", "nigeria": "ng",
    }
    return mapping.get(country)


def check_coords_entity_continent(rows: list[dict], **_kw) -> CheckResult:
    """Flag geographic outliers — assets on a different continent from the entity's majority.

    Uses reverse_geocoder if available, falls back to continent bounding boxes. Warn-only.
    """
    result = CheckResult("coords_entity_continent")

    rg = _try_reverse_geocoder()

    def _get_continent(lat: float, lon: float, cc: str = "") -> str | None:
        # Use country code if available (more accurate)
        if cc:
            cc = cc.upper()
            na = {"US", "CA", "MX", "GT", "HN", "SV", "NI", "CR", "PA", "CU", "DO", "HT", "JM", "TT", "BS", "PR"}
            sa = {"BR", "AR", "CL", "CO", "PE", "VE", "EC", "BO", "PY", "UY", "GY", "SR"}
            eu = {"GB", "DE", "FR", "IT", "ES", "NL", "BE", "AT", "CH", "SE", "NO", "DK", "FI",
                  "PL", "CZ", "PT", "IE", "GR", "HU", "RO", "BG", "HR", "SK", "SI", "LT", "LV", "EE"}
            af = {"ZA", "EG", "NG", "KE", "ET", "TZ", "GH", "MA", "TN", "DZ", "SN", "CI", "CM"}
            asia = {"CN", "JP", "KR", "IN", "TW", "SG", "TH", "MY", "ID", "PH", "VN",
                    "SA", "AE", "IL", "TR", "PK", "BD", "LK", "MM", "KH", "LA"}
            oc = {"AU", "NZ", "FJ", "PG", "WS", "TO"}
            if cc in na: return "North America"
            if cc in sa: return "South America"
            if cc in eu: return "Europe"
            if cc in af: return "Africa"
            if cc in asia: return "Asia"
            if cc in oc: return "Oceania"
        # Fallback: bounding box
        for cname, (min_lat, max_lat, min_lon, max_lon) in CONTINENT_BBOX.items():
            if min_lat <= lat <= max_lat and min_lon <= lon <= max_lon:
                return cname
        return None

    # Collect all coords for batch reverse geocoding
    geo_rows = []
    for i, row in enumerate(rows):
        entity = row.get("entity_name", "").strip()
        lat_s = row.get("latitude", "").strip()
        lon_s = row.get("longitude", "").strip()
        name = row.get("name", "").strip()
        if not entity or not lat_s or not lon_s:
            continue
        try:
            lat, lon = float(lat_s), float(lon_s)
        except ValueError:
            continue
        geo_rows.append((i, lat, lon, name, entity))

    if not geo_rows:
        return result

    # Batch reverse geocode
    ccs = [""] * len(geo_rows)
    if rg:
        coords = [(lat, lon) for _, lat, lon, _, _ in geo_rows]
        geo_results = rg.search(coords)
        ccs = [g["cc"] for g in geo_results]

    # Group by entity with continent
    entity_assets: dict[str, list[tuple[int, str, str]]] = {}  # entity -> [(idx, name, continent)]
    for (idx, lat, lon, name, entity), cc in zip(geo_rows, ccs):
        continent = _get_continent(lat, lon, cc)
        if continent:
            entity_assets.setdefault(entity, []).append((idx, name, continent))

    from collections import Counter
    for entity, assets in entity_assets.items():
        if len(assets) < 3:
            continue

        continent_counts = Counter(a[2] for a in assets)
        majority_continent, majority_count = continent_counts.most_common(1)[0]

        if majority_count / len(assets) < 0.6:
            continue

        for idx, name, continent in assets:
            if continent != majority_continent:
                result.warn(
                    f"Row {idx} '{name}' ({entity}): in {continent} but "
                    f"{majority_count}/{len(assets)} entity assets are in {majority_continent}"
                )

    return result


def check_entity_stake(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """entity_stake_pct is 0-100 when present. Fix fills empty with 100.0."""
    result = CheckResult("entity_stake_pct")
    empty_count = 0
    for i, row in enumerate(rows):
        val = row.get("entity_stake_pct", "").strip()
        if not val:
            if fix:
                rows[i]["entity_stake_pct"] = "100.0"
                empty_count += 1
            continue
        try:
            pct = float(val)
        except ValueError:
            result.fail(f"Row {i}: non-numeric entity_stake_pct '{val}'")
            continue
        if not (0 <= pct <= 100):
            result.fail(f"Row {i}: entity_stake_pct {pct} out of range [0, 100]")
    if empty_count:
        result.fix(f"Filled {empty_count} empty entity_stake_pct with 100.0")
    return result


def check_capacity_non_negative(rows: list[dict], **_kw) -> CheckResult:
    """Capacity values are non-negative when present."""
    result = CheckResult("capacity_non_negative")
    for i, row in enumerate(rows):
        val = row.get("capacity", "").strip()
        if not val:
            continue
        try:
            cap = float(val)
        except ValueError:
            result.fail(f"Row {i}: non-numeric capacity '{val}'")
            continue
        if cap < 0:
            result.fail(f"Row {i}: negative capacity {cap}")
    return result


def check_capacity_plausibility(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """Capacity values are plausible for their units.

    Deterministic detection flags implausible values. LLM fix reasons about
    asset type + name + value + units together to propose corrections — this
    is critical because capacity/units are tightly coupled and context-dependent.
    """
    result = CheckResult("capacity_plausibility")

    # Phase 1: collect implausible rows
    implausible: list[tuple[int, float, str, str, str]] = []  # (idx, cap, units, name, asset_type)
    for i, row in enumerate(rows):
        val_s = row.get("capacity", "").strip()
        units = row.get("capacity_units", "").strip()
        name = row.get("name", "").strip()
        if not val_s or not units:
            continue
        try:
            cap = float(val_s)
        except ValueError:
            continue
        if cap <= 0:
            continue

        bounds = CAPACITY_PLAUSIBILITY.get(units)
        if not bounds:
            continue

        lo, hi = bounds
        if lo <= cap <= hi:
            continue

        asset_type = row.get("asset_type_raw", "").strip()
        implausible.append((i, cap, units, name, asset_type))

    if not implausible:
        return result

    # Phase 2: report or fix
    if not fix_llm:
        # Deterministic hints for diagnostics
        for idx, cap, units, name, asset_type in implausible:
            bounds = CAPACITY_PLAUSIBILITY[units]
            hint = ""
            if units in AREA_UNIT_CONVERSIONS and cap < bounds[0]:
                candidates = []
                for other_unit in AREA_UNIT_CONVERSIONS:
                    if other_unit == units:
                        continue
                    ob = CAPACITY_PLAUSIBILITY.get(other_unit)
                    if ob and ob[0] <= cap <= ob[1]:
                        mid = math.sqrt(ob[0] * ob[1])
                        score = abs(math.log10(cap / mid))
                        converted = cap * AREA_UNIT_CONVERSIONS[other_unit] / AREA_UNIT_CONVERSIONS[units]
                        candidates.append((score, other_unit, converted))
                if candidates:
                    candidates.sort()
                    _, best_unit, best_converted = candidates[0]
                    hint = f" (looks like value is in {best_unit} → should be {best_converted:.1f} {units})"
            if cap < bounds[0]:
                result.fail(f"Row {idx} '{name}': capacity={cap} {units} is implausibly small{hint}")
            else:
                result.fail(f"Row {idx} '{name}': capacity={cap} {units} is implausibly large")
        return result

    # Phase 3: LLM fix — send all implausible rows in one batch
    llm = _try_llm_import()
    if not llm:
        result.fail("--fix-llm requires litellm: pip install ald-checker[llm]")
        return result

    try:
        corrections = llm.fix_capacity(
            [
                {"row": idx, "name": name, "asset_type": asset_type,
                 "capacity": cap, "capacity_units": units}
                for idx, cap, units, name, asset_type in implausible
            ],
            model=model or llm.DEFAULT_MODEL,
        )
        for c in corrections:
            idx = c["row"]
            new_cap = c.get("capacity")
            new_units = c.get("capacity_units")
            old_cap = rows[idx].get("capacity", "")
            old_units = rows[idx].get("capacity_units", "")
            name = rows[idx].get("name", "")
            if c.get("drop"):
                rows[idx]["capacity"] = ""
                rows[idx]["capacity_units"] = ""
                result.fix(f"Row {idx} '{name}': dropped implausible capacity {old_cap} {old_units}")
            elif new_cap is not None and new_units:
                rows[idx]["capacity"] = str(new_cap)
                rows[idx]["capacity_units"] = new_units
                result.fix(f"Row {idx} '{name}': {old_cap} {old_units} → {new_cap} {new_units}")
            else:
                result.fail(f"Row {idx} '{name}': capacity={old_cap} {old_units} is implausible — LLM could not resolve")
    except Exception as e:
        result.fail(f"LLM capacity fix failed: {e}")
        for idx, cap, units, name, _ in implausible:
            result.fail(f"Row {idx} '{name}': capacity={cap} {units} is implausible")

    return result


def check_capacity_units_appropriate(rows: list[dict], fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """Capacity units are semantically appropriate for the asset type and name.

    Catches things like a semiconductor fab measured in sqm instead of wafers/month,
    or an office measured in MW. Only runs with --fix-llm since this requires reasoning.
    """
    result = CheckResult("capacity_units_appropriate")

    if not fix_llm:
        return result

    # Collect rows that have both capacity and asset type
    to_check = []
    for i, row in enumerate(rows):
        cap = row.get("capacity", "").strip()
        units = row.get("capacity_units", "").strip()
        atype = row.get("asset_type_raw", "").strip()
        name = row.get("name", "").strip()
        if cap and units and atype:
            to_check.append({"row": i, "name": name, "asset_type": atype,
                             "capacity": cap, "capacity_units": units})

    if not to_check:
        return result

    llm = _try_llm_import()
    if not llm:
        return result

    try:
        flagged = llm.check_capacity_units_appropriate(to_check, model=model or llm.DEFAULT_MODEL)
        for f in flagged:
            idx = f["row"]
            name = rows[idx].get("name", "")
            old_cap = rows[idx].get("capacity", "")
            old_units = rows[idx].get("capacity_units", "")
            issue = f.get("issue", "")
            new_cap = f.get("capacity")
            new_units = f.get("capacity_units")

            if issue:
                result.warn(f"Row {idx} '{name}': capacity={old_cap} {old_units} — {issue}")
    except Exception as e:
        result.warn(f"LLM capacity-units-appropriate check failed: {e}")

    return result


def _extract_base_status(status: str) -> str:
    """Extract the base status from a compound status string.

    'operational (P1); under construction (P2-P3)' → 'operational'
    'under construction (production targeted 2028)' → 'under construction'
    'ramping (P1 mass production 1H 2026)' → 'ramping'
    """
    # Split on semicolon and take first segment, then strip parenthetical
    first_segment = status.split(";")[0].strip()
    base = re.split(r"\s*\(", first_segment)[0].strip()
    return base


def check_status_values(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """Status base value is one of the valid ALD statuses.

    Allows parenthetical detail (e.g. 'operational (P1-P2)') and compound
    statuses with semicolons (e.g. 'operational (P1); under construction (P2)').
    The base value (before first parenthetical) must be valid.
    """
    result = CheckResult("status_values")
    invalid: dict[str, list[int]] = {}
    needs_casing: list[int] = []
    for i, row in enumerate(rows):
        status = row.get("status", "").strip()
        if not status:
            continue
        base = _extract_base_status(status)
        base_lower = base.lower()
        if base_lower not in VALID_STATUSES:
            invalid.setdefault(status, []).append(i)
        elif base != base_lower:
            # Valid but wrong casing (e.g. "Temporarily closed" → "temporarily closed")
            needs_casing.append(i)

    # Fix casing on valid-but-miscased statuses
    if needs_casing and (fix or fix_llm):
        for idx in needs_casing:
            old = rows[idx]["status"]
            rows[idx]["status"] = old.lower()
        result.fix(f"Lowercased {len(needs_casing)} status values")
    elif needs_casing:
        result.fail(f"{len(needs_casing)} status values have wrong casing (e.g. '{rows[needs_casing[0]]['status']}')")

    if not invalid:
        return result

    if not fix and not fix_llm:
        for status, idxs in invalid.items():
            result.fail(f"Invalid status '{status}' at {len(idxs)} rows (first: {idxs[:5]})")
        return result

    still_invalid: dict[str, list[int]] = {}
    for status, idxs in invalid.items():
        base = _extract_base_status(status).lower()
        canonical = STATUS_ALIASES.get(base)
        if canonical:
            # Replace only the base portion, preserve parenthetical detail
            for idx in idxs:
                old_status = rows[idx]["status"]
                new_status = old_status.replace(
                    _extract_base_status(old_status), canonical, 1
                )
                rows[idx]["status"] = new_status
            result.fix(f"'{status}' base → '{canonical}' at {len(idxs)} rows")
        else:
            still_invalid[status] = idxs

    if still_invalid and fix_llm:
        llm = _try_llm_import()
        if llm:
            try:
                llm_map = llm.classify_status(list(still_invalid.keys()), model or llm.DEFAULT_MODEL)
                for status, idxs in still_invalid.items():
                    canonical = llm_map.get(status, "")
                    if canonical in VALID_STATUSES and canonical:
                        for idx in idxs:
                            rows[idx]["status"] = canonical
                        result.fix(f"LLM: '{status}' → '{canonical}' at {len(idxs)} rows")
                    else:
                        result.fail(f"Invalid status '{status}' at {len(idxs)} rows — LLM returned '{canonical}'")
            except Exception as e:
                result.fail(f"LLM status classification failed: {e}")
                for status, idxs in still_invalid.items():
                    result.fail(f"Invalid status '{status}' at {len(idxs)} rows (first: {idxs[:5]})")
        else:
            result.fail("--fix-llm requires litellm: pip install ald-checker[llm]")
    elif still_invalid:
        for status, idxs in still_invalid.items():
            result.fail(f"Invalid status '{status}' at {len(idxs)} rows (first: {idxs[:5]}) — no known alias")

    return result


def check_required_fields(rows: list[dict], **_kw) -> CheckResult:
    """entity_name and name (asset name) are never empty."""
    result = CheckResult("required_fields")
    missing_entity = [i for i, r in enumerate(rows) if not r.get("entity_name", "").strip()]
    missing_name = [i for i, r in enumerate(rows) if not r.get("name", "").strip()]
    if missing_entity:
        result.fail(f"Empty entity_name at {len(missing_entity)} rows: {missing_entity[:10]}")
    if missing_name:
        result.fail(f"Empty name (asset name) at {len(missing_name)} rows: {missing_name[:10]}")
    return result


def check_name_casing(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """Asset names are title-cased, not ALL CAPS or all lowercase."""
    result = CheckResult("name_casing")
    bad_rows = []
    for i, row in enumerate(rows):
        name = row.get("name", "").strip()
        if not name:
            continue
        if (name == name.upper() and len(name) > 3) or (name == name.lower() and len(name) > 1):
            bad_rows.append(i)

    if not bad_rows:
        return result

    if fix:
        for i in bad_rows:
            rows[i]["name"] = _smart_title_case(rows[i]["name"])
        result.fix(f"Title-cased {len(bad_rows)} asset names")
    else:
        samples = [rows[i]["name"] for i in bad_rows[:5]]
        result.fail(f"Bad casing at {len(bad_rows)} rows (samples: {samples})")
    return result


def check_entity_name_casing(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """Entity names are title-cased, not ALL CAPS or all lowercase."""
    result = CheckResult("entity_name_casing")
    entity_rows: dict[str, list[int]] = {}
    for i, row in enumerate(rows):
        name = row.get("entity_name", "").strip()
        if name:
            entity_rows.setdefault(name, []).append(i)

    bad_entities: dict[str, list[int]] = {}
    for name, idxs in entity_rows.items():
        if (name == name.upper() and len(name) > 3) or (name == name.lower() and len(name) > 1):
            bad_entities[name] = idxs

    if not bad_entities:
        return result

    if fix:
        for name, idxs in bad_entities.items():
            fixed_name = _smart_title_case(name)
            for i in idxs:
                rows[i]["entity_name"] = fixed_name
            result.fix(f"'{name}' → '{fixed_name}' at {len(idxs)} rows")
    else:
        for name, idxs in bad_entities.items():
            result.fail(f"Bad casing: '{name}' at {len(idxs)} rows")
    return result


def check_address_exists(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """Flag rows with coords but no address. Fix by reverse geocoding."""
    result = CheckResult("address_exists")
    missing = []
    for i, row in enumerate(rows):
        lat = row.get("latitude", "").strip() if isinstance(row.get("latitude"), str) else str(row.get("latitude", ""))
        addr = row.get("address", "").strip()
        if lat and lat not in ("", "None") and (not addr or addr == "None"):
            missing.append(i)

    if not missing:
        return result

    if fix and missing:
        # Check config — reverse geocoding can be disabled
        if not CONFIG.get("checks", {}).get("reverse_geocode", True):
            result.warn(f"{len(missing)} rows missing address (reverse_geocode disabled in config)")
            return result

        fixed = 0
        import urllib.request

        # Try Google reverse geocode first (returns English addresses)
        google_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")
        if google_key:
            for i in missing:
                row = rows[i]
                lat, lon = row.get("latitude"), row.get("longitude")
                try:
                    url = (f"https://maps.googleapis.com/maps/api/geocode/json"
                           f"?latlng={lat},{lon}&language=en&key={google_key}")
                    req = urllib.request.Request(url)
                    resp = urllib.request.urlopen(req, timeout=10)
                    data = json.loads(resp.read())
                    if data.get("results"):
                        addr = data["results"][0].get("formatted_address", "")
                        if addr:
                            rows[i]["address"] = addr
                            fixed += 1
                except Exception:
                    pass
        else:
            # Fallback to Nominatim (local language, rate limited)
            import time
            for i in missing[:50]:
                row = rows[i]
                lat, lon = row.get("latitude"), row.get("longitude")
                try:
                    url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json&accept-language=en"
                    req = urllib.request.Request(url, headers={"User-Agent": "ald-checker/1.0"})
                    resp = urllib.request.urlopen(req, timeout=10)
                    data = json.loads(resp.read())
                    addr = data.get("display_name", "")
                    if addr:
                        rows[i]["address"] = addr
                        fixed += 1
                    time.sleep(1.1)
                except Exception:
                    pass
        if fixed:
            result.fix(f"Reverse geocoded {fixed} addresses (of {len(missing)} missing)")
        remaining = len(missing) - fixed
        if remaining:
            result.warn(f"{remaining} rows still missing address")
    else:
        result.fail(f"{len(missing)} rows have coords but no address")
    return result


def check_supplementary_details(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """supplementary_details is valid JSON dict when present. Fix converts key:val format."""
    result = CheckResult("supplementary_details_json")
    fixed_count = 0
    for i, row in enumerate(rows):
        val = row.get("supplementary_details", "").strip()
        if not val:
            continue
        try:
            parsed = json.loads(val)
            if not isinstance(parsed, dict):
                result.fail(f"Row {i}: supplementary_details is {type(parsed).__name__}, expected dict")
        except json.JSONDecodeError:
            if fix:
                # Try to parse "key: val; key: val" format
                pairs = {}
                for part in val.split(";"):
                    part = part.strip()
                    if ":" in part:
                        k, v = part.split(":", 1)
                        pairs[k.strip()] = v.strip()
                    elif part:
                        pairs["info"] = part
                if pairs:
                    rows[i]["supplementary_details"] = json.dumps(pairs)
                    fixed_count += 1
                else:
                    result.fail(f"Row {i}: could not parse supplementary_details: {val[:80]}")
            else:
                result.fail(f"Row {i}: supplementary_details is not valid JSON: {val[:80]}...")
    if fixed_count:
        result.fix(f"Converted {fixed_count} key:val strings to JSON")
    return result


def check_entity_parent_consistency(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """Same entity_name always maps to same parent_name and parent_isin."""
    result = CheckResult("entity_parent_consistency")
    from collections import Counter
    entity_parents: dict[str, Counter] = {}
    for row in rows:
        en = row.get("entity_name", "").strip()
        pn = row.get("parent_name", "").strip()
        pi = row.get("parent_isin", "").strip()
        if en:
            key = f"{pn}||{pi}"
            entity_parents.setdefault(en, Counter())[key] += 1

    for entity, parents in entity_parents.items():
        if len(parents) <= 1:
            continue
        winner_key, _ = parents.most_common(1)[0]
        winner_pn, winner_pi = winner_key.split("||")
        minority = {k: v for k, v in parents.items() if k != winner_key}

        if fix:
            fixed = 0
            for row in rows:
                if row.get("entity_name", "").strip() == entity:
                    current = f"{row.get('parent_name', '').strip()}||{row.get('parent_isin', '').strip()}"
                    if current != winner_key:
                        row["parent_name"] = winner_pn
                        row["parent_isin"] = winner_pi
                        fixed += 1
            if fixed:
                result.fix(f"'{entity}': standardized parent to '{winner_pn}' at {fixed} rows")
        else:
            result.fail(f"'{entity}' has multiple parents: {dict(parents)}")
    return result


def check_entity_isin_valid(rows: list[dict], **_kw) -> CheckResult:
    """Validate entity_isin exists in corp-graph if present."""
    result = CheckResult("entity_isin_valid")
    isins = set()
    for row in rows:
        isin = row.get("entity_isin", "").strip()
        if isin:
            isins.add(isin)

    if not isins:
        return result

    # Check against corp-graph
    try:
        import psycopg
        from psycopg.rows import dict_row
        import os
        db_url = os.environ.get("CORPGRAPH_DB_URL", "postgresql://corpgraph:corpgraph@localhost:5432/corpgraph")
        conn = psycopg.connect(db_url, row_factory=dict_row)
        for isin in isins:
            cur = conn.execute(
                "SELECT COUNT(*) as cnt FROM company_universe WHERE %s = ANY(isin_list)", (isin,)
            )
            if cur.fetchone()["cnt"] == 0:
                result.fail(f"ISIN '{isin}' not found in corp-graph")
        conn.close()
    except Exception:
        pass  # Skip if corp-graph not available
    return result


def check_duplicate_assets(rows: list[dict], **_kw) -> CheckResult:
    """Flag assets that share same name + same entity + identical/near-identical coords + similar address."""
    result = CheckResult("duplicate_assets")

    # Group by name + entity
    key_rows: dict[str, list[int]] = {}
    for i, row in enumerate(rows):
        name = row.get("name", "").strip().lower()
        entity = row.get("entity_name", "").strip().lower()
        if name and entity:
            key = f"{entity}||{name}"
            key_rows.setdefault(key, []).append(i)

    for key, idxs in key_rows.items():
        if len(idxs) <= 1:
            continue

        # Same name + entity — now check if coords are near-identical
        for a_idx in range(len(idxs)):
            for b_idx in range(a_idx + 1, len(idxs)):
                row_a = rows[idxs[a_idx]]
                row_b = rows[idxs[b_idx]]
                try:
                    lat_a, lon_a = float(row_a["latitude"]), float(row_a["longitude"])
                    lat_b, lon_b = float(row_b["latitude"]), float(row_b["longitude"])
                    dist = math.sqrt((lat_a - lat_b) ** 2 + (lon_a - lon_b) ** 2) * 111_000
                    if dist < 100:  # Within ~100m
                        entity, name = key.split("||")
                        result.fail(
                            f"Duplicate: '{name}' for '{entity}' at rows {idxs[a_idx]} and {idxs[b_idx]} ({dist:.0f}m apart)"
                        )
                except (ValueError, TypeError, KeyError):
                    # If coords missing, fall back to same name = duplicate
                    entity, name = key.split("||")
                    result.fail(f"Duplicate asset name '{name}' for '{entity}' at rows {idxs}")
                    break

    return result


def check_date_researched(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """date_researched is a valid YYYY-MM-DD date."""
    result = CheckResult("date_researched")
    date_re = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    today = date.today().isoformat()

    empty = [i for i, r in enumerate(rows) if not r.get("date_researched", "").strip()]
    invalid: dict[str, list[int]] = {}
    for i, r in enumerate(rows):
        val = r.get("date_researched", "").strip()
        if val and not date_re.match(val):
            invalid.setdefault(val, []).append(i)

    if fix and empty:
        for i in empty:
            rows[i]["date_researched"] = today
        result.fix(f"Filled empty date_researched with '{today}' at {len(empty)} rows")
    elif empty:
        result.fail(f"Empty date_researched at {len(empty)} rows: {empty[:10]}")

    if invalid:
        # Try common deterministic formats first
        from datetime import datetime
        FORMATS = ["%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%B %d, %Y", "%b %d, %Y",
                    "%d %B %Y", "%d %b %Y", "%m-%d-%Y", "%Y%m%d"]
        still_invalid: dict[str, list[int]] = {}
        for val, idxs in invalid.items():
            parsed = None
            for fmt in FORMATS:
                try:
                    parsed = datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
            if parsed:
                if fix:
                    for i in idxs:
                        rows[i]["date_researched"] = parsed
                    result.fix(f"Date: '{val}' → '{parsed}' at {len(idxs)} rows")
            else:
                still_invalid[val] = idxs

        # LLM fallback for weird formats
        if still_invalid and fix_llm:
            llm = _try_llm_import()
            if llm:
                try:
                    parsed_map = llm.parse_dates(list(still_invalid.keys()), model=model or llm.DEFAULT_MODEL)
                    for orig, parsed in parsed_map.items():
                        if orig in still_invalid and date_re.match(parsed):
                            for i in still_invalid[orig]:
                                rows[i]["date_researched"] = parsed
                            result.fix(f"LLM date: '{orig}' → '{parsed}' at {len(still_invalid[orig])} rows")
                except Exception:
                    pass

        for val, idxs in still_invalid.items():
            if not any(val in msg for msg in result.fixed):
                result.fail(f"Invalid date format '{val}' at {len(idxs)} rows")
    return result


def check_isin_format(rows: list[dict], **_kw) -> CheckResult:
    """ISINs are 12 alphanumeric characters when present."""
    result = CheckResult("isin_format")
    isin_re = re.compile(r"^[A-Z]{2}[A-Z0-9]{10}$")
    for i, row in enumerate(rows):
        for col in ("entity_isin", "parent_isin"):
            val = row.get(col, "").strip()
            if val and not isin_re.match(val):
                result.fail(f"Row {i}: {col} '{val}' is not a valid ISIN (expected 2 letters + 10 alphanumeric)")
    return result


def check_capacity_units_consistency(rows: list[dict], fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """Same asset_type_raw has consistent capacity_units.

    Warns on mixed units. LLM fix converts BOTH the value and units together —
    never change units without converting the value.
    """
    result = CheckResult("capacity_units_consistency")
    mapping: dict[str, dict[str, list[int]]] = {}
    for i, row in enumerate(rows):
        raw = row.get("asset_type_raw", "").strip().lower()
        units = row.get("capacity_units", "").strip()
        cap = row.get("capacity", "").strip()
        if raw and units and cap:
            mapping.setdefault(raw, {}).setdefault(units, []).append(i)

    mixed = {}
    for raw, units_map in mapping.items():
        if len(units_map) > 1:
            counts = {u: len(idxs) for u, idxs in units_map.items()}
            result.warn(f"'{raw}' has mixed capacity_units: {counts}")
            mixed[raw] = units_map

    if not mixed or not fix_llm:
        return result

    # Collect minority-unit rows that need conversion
    to_fix = []
    for raw, units_map in mixed.items():
        majority_unit = max(units_map, key=lambda u: len(units_map[u]))
        for unit, idxs in units_map.items():
            if unit == majority_unit:
                continue
            for idx in idxs:
                to_fix.append({
                    "row": idx,
                    "name": rows[idx].get("name", ""),
                    "asset_type": raw,
                    "capacity": rows[idx].get("capacity", ""),
                    "capacity_units": unit,
                    "target_units": majority_unit,
                })

    if not to_fix:
        return result

    llm = _try_llm_import()
    if not llm:
        return result

    try:
        corrections = llm.convert_capacity_units(to_fix, model=model or llm.DEFAULT_MODEL)
        for c in corrections:
            idx = c["row"]
            old_cap = rows[idx].get("capacity", "")
            old_units = rows[idx].get("capacity_units", "")
            new_cap = c.get("capacity")
            new_units = c.get("capacity_units")
            name = rows[idx].get("name", "")
            if new_cap is not None and new_units:
                rows[idx]["capacity"] = str(new_cap)
                rows[idx]["capacity_units"] = new_units
                result.fix(f"Row {idx} '{name}': {old_cap} {old_units} → {new_cap} {new_units}")
    except Exception as e:
        result.warn(f"LLM capacity unit conversion failed: {e}")

    return result


def check_source_url_format(rows: list[dict], **_kw) -> CheckResult:
    """Source URLs are valid http(s) URLs when present."""
    result = CheckResult("source_url_format")
    url_re = re.compile(r"^https?://\S+$")
    for i, row in enumerate(rows):
        url = row.get("source_url", "").strip()
        if url and not url_re.match(url):
            result.fail(f"Row {i}: invalid source_url '{url[:80]}'")
    return result


def check_entity_name_consistency(rows: list[dict], fix: bool = False, **_kw) -> CheckResult:
    """Entity names that resolve to the same base name are written the same way.

    Uses cleanco to strip legal suffixes (Inc., Ltd., GmbH, A/S, etc.)
    and groups by base name. Majority form wins on --fix.
    """
    from cleanco import basename

    result = CheckResult("entity_name_consistency")
    base_to_forms: dict[str, dict[str, list[int]]] = {}
    for i, row in enumerate(rows):
        name = row.get("entity_name", "").strip()
        if not name:
            continue
        base = basename(name).strip().lower()
        if not base:
            continue
        base_to_forms.setdefault(base, {}).setdefault(name, []).append(i)

    for base, forms in base_to_forms.items():
        if len(forms) <= 1:
            continue
        counts = {form: len(idxs) for form, idxs in forms.items()}
        if fix:
            winner = max(forms, key=lambda f: len(forms[f]))
            fixed_count = 0
            for form, idxs in forms.items():
                if form != winner:
                    for idx in idxs:
                        rows[idx]["entity_name"] = winner
                    fixed_count += len(idxs)
            if fixed_count:
                result.fix(f"Normalized to '{winner}' (was {counts}, fixed {fixed_count} rows)")
        else:
            result.fail(f"Same entity written as: {counts}")
    return result


def check_coordinate_proximity(rows: list[dict], **_kw) -> CheckResult:
    """Flag asset pairs within proximity threshold of each other."""
    result = CheckResult("coordinate_proximity")
    THRESHOLD_M = CONFIG.get("thresholds", {}).get("proximity_m", 25)

    geo_rows: list[tuple[int, float, float, str]] = []
    for i, row in enumerate(rows):
        lat_s = row.get("latitude", "").strip()
        lon_s = row.get("longitude", "").strip()
        if not lat_s or not lon_s:
            continue
        try:
            lat, lon = float(lat_s), float(lon_s)
            name = row.get("name", "").strip()
            geo_rows.append((i, lat, lon, name))
        except ValueError:
            continue

    def _dist_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        R = 6_371_000
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = (math.sin(dlat / 2) ** 2
             + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
             * math.sin(dlon / 2) ** 2)
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    DEG_THRESHOLD = THRESHOLD_M / 111_000 * 1.5

    for a in range(len(geo_rows)):
        idx_a, lat_a, lon_a, name_a = geo_rows[a]
        for b in range(a + 1, len(geo_rows)):
            idx_b, lat_b, lon_b, name_b = geo_rows[b]
            if abs(lat_a - lat_b) > DEG_THRESHOLD or abs(lon_a - lon_b) > DEG_THRESHOLD:
                continue
            dist = _dist_m(lat_a, lon_a, lat_b, lon_b)
            if dist < THRESHOLD_M:
                result.warn(
                    f"Rows {idx_a} & {idx_b} are {dist:.0f}m apart: "
                    f"'{name_a}' vs '{name_b}'"
                )
    return result


def check_attribution_source(rows: list[dict], fix: bool = False, fix_llm: bool = False, model: str = "", **_kw) -> CheckResult:
    """attribution_source is set and standardized on every row."""
    result = CheckResult("attribution_source")
    empty = [i for i, r in enumerate(rows) if not r.get("attribution_source", "").strip()]
    if fix and empty:
        for i in empty:
            rows[i]["attribution_source"] = "asset_discovery"
        result.fix(f"Filled empty attribution_source with 'asset_discovery' at {len(empty)} rows")
    elif empty:
        result.fail(f"Empty attribution_source at {len(empty)} rows: {empty[:10]}")

    # LLM standardize messy values (only if they look like sentences/descriptions, not clean identifiers)
    if fix_llm:
        nonstandard: dict[str, list[int]] = {}
        for i, r in enumerate(rows):
            src = r.get("attribution_source", "").strip()
            if src and (" " in src or src[0].isupper()):  # Looks like free text, not a clean identifier
                nonstandard.setdefault(src, []).append(i)
        if nonstandard:
            llm = _try_llm_import()
            if llm:
                try:
                    mapped = llm.standardize_attribution(list(nonstandard.keys()), model=model or llm.DEFAULT_MODEL)
                    for orig, std in mapped.items():
                        if orig in nonstandard and std != orig:
                            for idx in nonstandard[orig]:
                                rows[idx]["attribution_source"] = std
                            result.fix(f"Attribution: '{orig}' → '{std}' at {len(nonstandard[orig])} rows")
                except Exception:
                    pass
    return result


# ── Check registry + runner ──────────────────────────────────────────────────

ALL_CHECKS = [
    # Phase 1: Column structure
    check_columns,
    check_none_strings,
    check_numeric_cleanup,
    check_asset_id_unique,
    # Phase 2: Classification (order matters: raw type → NS → GICS)
    check_asset_type_raw_standardize,
    check_naturesense_correct,
    check_gics_correct,
    check_naturesense_valid,
    check_naturesense_consistency,
    check_gics_valid,
    check_gics_consistency,
    # Phase 3: Data quality
    check_coordinates,
    check_coords_country_match,
    check_coords_entity_continent,
    check_address_exists,
    check_entity_stake,
    check_capacity_non_negative,
    check_capacity_plausibility,
    check_capacity_units_appropriate,
    check_status_values,
    check_required_fields,
    check_name_casing,
    check_entity_name_casing,
    check_supplementary_details,
    check_date_researched,
    check_isin_format,
    check_capacity_units_consistency,
    check_source_url_format,
    check_attribution_source,
    # Phase 4: Entity structure
    check_entity_name_consistency,
    check_entity_parent_consistency,
    check_entity_isin_valid,
    # Phase 5: Duplicates & proximity
    check_duplicate_assets,
    check_coordinate_proximity,
]


def run_checks(
    csv_path: str,
    fix: bool = False,
    fix_llm: bool = False,
    model: str = "",
    only_checks: list[str] | None = None,
    skip_checks: list[str] | None = None,
    no_xlsx: bool = False,
    dry_run: bool = False,
) -> list[CheckResult]:
    """Run all checks on a CSV file. Returns list of CheckResults."""
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"{path} not found")

    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        rows = list(reader)

    # Apply config defaults
    if not model:
        model = CONFIG.get("llm", {}).get("model", "openai/gpt-4.1-nano")

    # Merge config skip checks with CLI skip
    config_skip = CONFIG.get("checks", {}).get("skip", [])
    if config_skip:
        skip_checks = list(set((skip_checks or []) + config_skip))

    print(f"\n{'='*60}")
    print(f"Checking: {path}")
    print(f"Rows: {len(rows)}")
    if fix_llm:
        print(f"LLM model: {model}")
    if only_checks:
        print(f"Only: {only_checks}")
    if skip_checks:
        print(f"Skip: {skip_checks}")
    if dry_run:
        print("DRY RUN — no output files will be written")
    print(f"{'='*60}")

    # Filter checks
    checks_to_run = ALL_CHECKS
    if only_checks:
        checks_to_run = [fn for fn in ALL_CHECKS if fn.__name__.replace("check_", "") in only_checks]
    if skip_checks:
        checks_to_run = [fn for fn in checks_to_run if fn.__name__.replace("check_", "") not in skip_checks]

    results = []
    for check_fn in checks_to_run:
        if check_fn == check_columns:
            r = check_fn(rows, headers, fix=fix, fix_llm=fix_llm, model=model)
        else:
            r = check_fn(rows, fix=fix, fix_llm=fix_llm, model=model)
        results.append(r)

        if r.fixed:
            status = "\033[33mFIXED\033[0m"
        elif r.warnings and r.passed:
            status = "\033[33mWARN\033[0m"
        elif r.passed:
            status = "\033[32mPASS\033[0m"
        else:
            status = "\033[31mFAIL\033[0m"
        print(f"  [{status}] {r.name}")
        for issue in r.issues:
            print(f"         {issue}")
        for fixed_msg in r.fixed:
            print(f"         \033[33m{fixed_msg}\033[0m")
        for warn_msg in r.warnings:
            print(f"         \033[33m⚠ {warn_msg}\033[0m")

    # Write fixed output
    if (fix or fix_llm) and any(r.fixed for r in results) and not dry_run:
        # Sort rows: entity_name → asset_type_raw → name
        rows.sort(key=lambda r: (
            r.get("entity_name", "").lower(),
            r.get("asset_type_raw", "").lower(),
            r.get("name", "").lower(),
        ))

        # Write CSV — use headers which may have been updated by check_columns
        # Also include any extra columns that were added during fixes
        all_keys = set()
        for row in rows:
            all_keys.update(row.keys())
        out_headers = [h for h in headers if h in all_keys]
        for k in all_keys:
            if k not in out_headers and k:
                out_headers.append(k)

        suffix = CONFIG.get("output", {}).get("suffix", "_checked")
        csv_out = path.with_stem(path.stem + suffix)
        with csv_out.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=out_headers, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        print(f"\n  Fixed CSV written to: {csv_out}")

        # Write xlsx with Key + Assets + Review tabs
        try:
            if no_xlsx:
                raise ImportError("xlsx disabled")
            import openpyxl
            from collections import Counter

            wb = openpyxl.Workbook()

            # --- Key sheet ---
            ws_key = wb.active
            ws_key.title = "Key"
            # Find the top-level entity: most common parent_name, or most common entity_name if no parents
            from collections import Counter as _Counter
            parent_counts = _Counter(r.get("parent_name", "").strip() for r in rows if r.get("parent_name", "").strip())
            entity_counts = _Counter(r.get("entity_name", "").strip() for r in rows)
            if parent_counts:
                # Has parent/child structure — top entity is the most common parent
                top_entity = parent_counts.most_common(1)[0][0]
                # Find its ISIN from any row where it's entity_name
                top_isin = ""
                for r in rows:
                    if r.get("entity_name", "").strip() == top_entity and r.get("entity_isin", "").strip():
                        top_isin = r["entity_isin"].strip()
                        break
                # If not found as entity, check parent_isin
                if not top_isin:
                    for r in rows:
                        if r.get("parent_name", "").strip() == top_entity and r.get("parent_isin", "").strip():
                            top_isin = r["parent_isin"].strip()
                            break
                entity = top_entity
                isin = top_isin
                parent = ""
            else:
                entity = entity_counts.most_common(1)[0][0] if entity_counts else ""
                isin = ""
                for r in rows:
                    if r.get("entity_name", "").strip() == entity and r.get("entity_isin", "").strip():
                        isin = r["entity_isin"].strip()
                        break
                parent = ""

            r = 1
            ws_key.cell(r, 1, entity); ws_key.cell(r, 2, "ALD Output"); r += 2
            ws_key.cell(r, 1, "ISIN"); ws_key.cell(r, 2, isin or "N/A"); r += 1
            ws_key.cell(r, 1, "Entity"); ws_key.cell(r, 2, entity); r += 1
            if parent:
                ws_key.cell(r, 1, "Parent"); ws_key.cell(r, 2, parent); r += 1
            ws_key.cell(r, 1, "Total Assets"); ws_key.cell(r, 2, len(rows)); r += 2

            ns_counts = Counter(row.get("naturesense_asset_type", "") for row in rows)
            ws_key.cell(r, 1, "NatureSense Types"); r += 1
            for ns, cnt in ns_counts.most_common():
                ws_key.cell(r, 1, f"  {ns}"); ws_key.cell(r, 2, cnt); r += 1
            r += 1

            gics_counts = Counter(str(row.get("industry_code", "")) for row in rows)
            ws_key.cell(r, 1, "GICS Codes"); r += 1
            for gics, cnt in gics_counts.most_common():
                ws_key.cell(r, 1, f"  {gics}"); ws_key.cell(r, 2, cnt); r += 1
            r += 1

            type_counts = Counter(row.get("asset_type_raw", "") for row in rows)
            ws_key.cell(r, 1, "Asset Types"); r += 1
            for t, cnt in type_counts.most_common():
                ws_key.cell(r, 1, f"  {t}"); ws_key.cell(r, 2, cnt); r += 1
            r += 1

            ws_key.column_dimensions["A"].width = 50
            ws_key.column_dimensions["B"].width = 60

            # --- Audit sheet ---
            ws_audit = wb.create_sheet("Audit")
            ws_audit.cell(1, 1, "Check"); ws_audit.cell(1, 2, "Status"); ws_audit.cell(1, 3, "Detail")
            ar = 2
            for check_r in results:
                for msg in check_r.fixed:
                    ws_audit.cell(ar, 1, check_r.name); ws_audit.cell(ar, 2, "FIXED"); ws_audit.cell(ar, 3, msg); ar += 1
                for msg in check_r.issues:
                    ws_audit.cell(ar, 1, check_r.name); ws_audit.cell(ar, 2, "FAIL"); ws_audit.cell(ar, 3, msg); ar += 1
                for msg in check_r.warnings:
                    ws_audit.cell(ar, 1, check_r.name); ws_audit.cell(ar, 2, "WARN"); ws_audit.cell(ar, 3, msg); ar += 1
            ws_audit.column_dimensions["A"].width = 35
            ws_audit.column_dimensions["B"].width = 10
            ws_audit.column_dimensions["C"].width = 80

            # --- Assets sheet ---
            ws_assets = wb.create_sheet("Assets")
            out_cols = [h for h in ALD_COLUMNS if h in headers]
            # Add extra cols that exist
            for h in headers:
                if h in EXTRA_COLUMNS and h not in out_cols:
                    out_cols.append(h)

            for j, col in enumerate(out_cols, 1):
                ws_assets.cell(1, j, col)
            for i, row in enumerate(rows, 2):
                for j, col in enumerate(out_cols, 1):
                    ws_assets.cell(i, j, row.get(col, ""))

            xlsx_out = path.with_suffix(".xlsx")
            wb.save(str(xlsx_out))
            print(f"  Fixed XLSX written to: {xlsx_out}")
        except ImportError:
            print("  (openpyxl not installed — skipping xlsx output)")

    passed = sum(1 for r in results if r.passed and not r.fixed)
    fixed = sum(1 for r in results if r.fixed)
    failed = sum(1 for r in results if not r.passed and not r.fixed)
    print(f"\n  Summary: {passed} passed, {fixed} fixed, {failed} failed out of {len(results)} checks")

    return results
